from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from app.core.database import get_db
from app.models.asset import MFITAsset
from app.models.changelog import ChangeLog
from app.models.user_history import UserHistory
from app.api.auth import get_current_user
from app.models.auth import User
from app.core.depreciation import calculate_depreciation
from datetime import datetime, timezone
import io
import json
import csv

router = APIRouter(prefix="/assets", tags=["Assets"])


@router.get("/")
def list_assets(
    search: str = "",
    status: str = "",
    type: str = "",
    db: Session = Depends(get_db),
):
    """Return all assets, optionally filtered by search, status, or type."""
    query = db.query(MFITAsset)

    if search:
        pattern = f"%{search}%"
        query = query.filter(
            MFITAsset.tracking_code.ilike(pattern)
            | MFITAsset.users_name.ilike(pattern)
            | MFITAsset.brand.ilike(pattern)
            | MFITAsset.model.ilike(pattern)
            | MFITAsset.serial_no.ilike(pattern)
            | MFITAsset.department.ilike(pattern)
        )

    if status:
        query = query.filter(MFITAsset.assignment_status == status)

    if type:
        query = query.filter(MFITAsset.type == type)

    assets = query.order_by(MFITAsset.id).all()

    return [
        {
            "id": a.id,
            "no": a.no,
            "type": a.type,
            "tracking_code": a.tracking_code,
            "brand": a.brand,
            "model": a.model,
            "serial_no": a.serial_no,
            "assignment_status": a.assignment_status,
            "users_name": a.users_name,
            "department": a.department,
        }
        for a in assets
    ]


@router.get("/{asset_id}")
def get_asset(asset_id: int, db: Session = Depends(get_db)):
    """Return full details for a single asset by its database ID, including depreciation."""
    asset = db.query(MFITAsset).filter(MFITAsset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    data = {c.name: getattr(asset, c.name) for c in MFITAsset.__table__.columns}

    # Add depreciation info
    depreciation = calculate_depreciation(
        purchase_price=asset.price,
        purchase_date_str=asset.purchase_date,
        lifespan_years=5,
    )
    data["depreciation"] = depreciation

    return data


@router.post("/")
def create_asset(
    data: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Create a new asset. Returns the created asset with its ID."""
    allowed_fields = {c.name for c in MFITAsset.__table__.columns} - {"id"}
    asset_data = {k: v for k, v in data.items() if k in allowed_fields and v}

    asset = MFITAsset(**asset_data)
    db.add(asset)
    db.commit()
    db.refresh(asset)

    # Log creation
    log_entry = ChangeLog(
        asset_id=asset.id,
        field_name="__created__",
        old_value=None,
        new_value="Asset created",
        changed_by=user.display_name,
    )
    db.add(log_entry)

    # If a user is assigned, create user history
    if asset.users_name:
        history = UserHistory(
            asset_id=asset.id,
            user_name=asset.users_name,
            department=asset.department,
            assigned_by=user.display_name,
        )
        db.add(history)

    db.commit()

    return {
        "message": "Asset created",
        "id": asset.id,
        "tracking_code": asset.tracking_code,
    }


@router.put("/{asset_id}")
def update_asset(
    asset_id: int,
    updates: dict = Body(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Update asset fields. Automatically logs all changes to the change_log table.
    If users_name changes, also records user history.
    """
    asset = db.query(MFITAsset).filter(MFITAsset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    # Fields that can be updated
    allowed_fields = {c.name for c in MFITAsset.__table__.columns} - {"id"}
    old_user = asset.users_name
    old_dept = asset.department

    changes_made = []

    for field, new_value in updates.items():
        if field not in allowed_fields:
            continue

        old_value = getattr(asset, field)
        old_str = str(old_value) if old_value is not None else None
        new_str = str(new_value) if new_value is not None else None

        if old_str != new_str:
            # Log the change
            log_entry = ChangeLog(
                asset_id=asset_id,
                field_name=field,
                old_value=old_str,
                new_value=new_str,
                changed_by=user.display_name,
            )
            db.add(log_entry)
            changes_made.append({"field": field, "old": old_str, "new": new_str})

            # Apply the change
            setattr(asset, field, new_value)

    # Handle user history if user assignment changed
    new_user = asset.users_name
    if old_user != new_user:
        # Close old assignment
        if old_user:
            old_history = (
                db.query(UserHistory)
                .filter(
                    UserHistory.asset_id == asset_id,
                    UserHistory.user_name == old_user,
                    UserHistory.returned_date.is_(None),
                )
                .first()
            )
            if old_history:
                old_history.returned_date = datetime.now(timezone.utc)

        # Create new assignment
        if new_user:
            new_history = UserHistory(
                asset_id=asset_id,
                user_name=new_user,
                department=asset.department,
                assigned_by=user.display_name,
            )
            db.add(new_history)

    db.commit()

    return {
        "message": f"{len(changes_made)} field(s) updated",
        "changes": changes_made,
    }


@router.delete("/{asset_id}")
def delete_asset(
    asset_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Delete an asset and its related records."""
    asset = db.query(MFITAsset).filter(MFITAsset.id == asset_id).first()
    if not asset:
        raise HTTPException(status_code=404, detail="Asset not found")

    tracking = asset.tracking_code or f"ID:{asset_id}"

    # Delete related records
    db.query(ChangeLog).filter(ChangeLog.asset_id == asset_id).delete()
    db.query(UserHistory).filter(UserHistory.asset_id == asset_id).delete()

    # Import here to avoid circular imports
    from app.models.attachment import Attachment
    import os
    attachments = db.query(Attachment).filter(Attachment.asset_id == asset_id).all()
    upload_dir = os.path.join(os.path.dirname(__file__), "..", "..", "uploads", str(asset_id))
    for att in attachments:
        fpath = os.path.join(upload_dir, att.filename)
        if os.path.exists(fpath):
            os.remove(fpath)
    db.query(Attachment).filter(Attachment.asset_id == asset_id).delete()

    # Remove upload directory
    if os.path.exists(upload_dir):
        try:
            os.rmdir(upload_dir)
        except OSError:
            pass

    db.delete(asset)
    db.commit()

    return {"message": f"Asset '{tracking}' deleted"}


# ═══════════════════════════════════════════════════════════
# Import / Export endpoints
# ═══════════════════════════════════════════════════════════

# Column mapping: Excel header → DB column name
COLUMN_MAP = {
    "No": "no",
    "Type": "type",
    "Owner": "owner",
    "Tracking Code": "tracking_code",
    "Brand": "brand",
    "Model": "model",
    "Serial No": "serial_no",
    "CPU": "cpu",
    "RAM": "ram",
    "Storage": "storage",
    "Printer Type": "printer_type",
    "Printer Color": "printer_color",
    "Connectivity": "connectivity",
    "Function": "function",
    "Monitor Size": "monitor_size",
    "Input Type": "input_type",
    "Price": "price",
    "Purchase Date": "purchase_date",
    "Estimate Lifespan": "estimate_lifespan",
    "Expiry Date": "expiry_date",
    "Start Date": "start_date",
    "Used Years": "used_years",
    "End Date": "end_date",
    "Assignment Status": "assignment_status",
    "User's Name": "users_name",
    "Department": "department",
    "Assignment Date": "assignment_date",
}

REVERSE_COLUMN_MAP = {v: k for k, v in COLUMN_MAP.items()}


@router.post("/import")
async def import_excel(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Import assets from an Excel (.xlsx) file.
    Rows with a tracking_code that already exists are SKIPPED.
    Returns count of imported and skipped rows.
    """
    import openpyxl

    content = await file.read()
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    # First row = headers
    raw_headers = [str(h).strip() if h else "" for h in rows[0]]

    # Map headers to DB column names
    header_to_col = {}
    for i, h in enumerate(raw_headers):
        if h in COLUMN_MAP:
            header_to_col[i] = COLUMN_MAP[h]
        else:
            # Try case-insensitive match
            for excel_name, db_col in COLUMN_MAP.items():
                if h.lower() == excel_name.lower():
                    header_to_col[i] = db_col
                    break

    if not header_to_col:
        raise HTTPException(
            status_code=400,
            detail="No matching columns found. Expected headers like: " + ", ".join(list(COLUMN_MAP.keys())[:5]) + "..."
        )

    # Get all existing tracking codes for duplicate check
    existing_codes = set()
    for code in db.query(MFITAsset.tracking_code).all():
        if code[0]:
            existing_codes.add(str(code[0]).strip())

    imported = 0
    skipped = 0
    errors = []

    for row_idx, row in enumerate(rows[1:], start=2):
        try:
            asset_data = {}
            for col_idx, db_col in header_to_col.items():
                if col_idx < len(row):
                    val = row[col_idx]
                    if val is not None:
                        # Convert datetime objects to string
                        if hasattr(val, 'strftime'):
                            val = val.strftime("%Y-%m-%d %H:%M:%S")
                        asset_data[db_col] = str(val).strip()

            # Skip if no data
            if not asset_data:
                continue

            # Skip if tracking code already exists
            tc = asset_data.get("tracking_code", "").strip()
            if tc and tc in existing_codes:
                skipped += 1
                continue

            asset = MFITAsset(**asset_data)
            db.add(asset)
            db.flush()  # Get the ID

            # Log creation
            log_entry = ChangeLog(
                asset_id=asset.id,
                field_name="__imported__",
                old_value=None,
                new_value=f"Imported from {file.filename}",
                changed_by=user.display_name,
            )
            db.add(log_entry)

            if tc:
                existing_codes.add(tc)
            imported += 1

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)}")
            continue

    db.commit()
    wb.close()

    return {
        "message": f"Import complete: {imported} added, {skipped} skipped (duplicates)",
        "imported": imported,
        "skipped": skipped,
        "errors": errors[:10],  # Only return first 10 errors
    }


@router.get("/export/excel")
def export_excel(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export all assets to an Excel (.xlsx) file."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    assets = db.query(MFITAsset).order_by(MFITAsset.id).all()
    columns = [c.name for c in MFITAsset.__table__.columns if c.name != "id"]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MF IT Assets"

    # Header row with styling
    header_fill = PatternFill(start_color="6C5CE7", end_color="6C5CE7", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    headers = [REVERSE_COLUMN_MAP.get(c, c) for c in columns]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Data rows
    for row_idx, asset in enumerate(assets, 2):
        for col_idx, col_name in enumerate(columns, 1):
            val = getattr(asset, col_name)
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"MF_IT_Assets_{timestamp}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/json")
def export_json(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export all assets as a JSON file."""
    assets = db.query(MFITAsset).order_by(MFITAsset.id).all()
    data = []
    for a in assets:
        row = {c.name: getattr(a, c.name) for c in MFITAsset.__table__.columns}
        data.append(row)

    content = json.dumps(data, indent=2, default=str)
    output = io.BytesIO(content.encode("utf-8"))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"MF_IT_Assets_{timestamp}.json"

    return StreamingResponse(
        output,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/export/csv")
def export_csv(
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Export all assets as a CSV file."""
    assets = db.query(MFITAsset).order_by(MFITAsset.id).all()
    columns = [c.name for c in MFITAsset.__table__.columns if c.name != "id"]

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row with friendly names
    writer.writerow([REVERSE_COLUMN_MAP.get(c, c) for c in columns])

    # Data rows
    for a in assets:
        writer.writerow([getattr(a, c) or "" for c in columns])

    content = output.getvalue().encode("utf-8")
    buffer = io.BytesIO(content)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"MF_IT_Assets_{timestamp}.csv"

    return StreamingResponse(
        buffer,
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )